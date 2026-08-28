"""Çıktı: sınıf / öğretmen bazında HTML, PDF ve Excel."""
from __future__ import annotations

import io
from collections import defaultdict

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.db import get_db
from app.deps import aktif_donem, current_user
from app.models import Day, Institution, Term, Timetable
from app.routers.timetables import izgara_hucreleri

router = APIRouter(prefix="/timetables/{timetable_id}/export", tags=["çıktı"],
                   dependencies=[Depends(current_user)])

BAKIS = {"sube": "Şube", "ogretmen": "Öğretmen"}
DUZEN = {"ayri": "Ayrı sayfalar", "carsaf": "Çarşaf liste"}


def _izgara_yapisi(db: Session, donem: Term) -> tuple[list[Day], list[int]]:
    """Dönemin aktif günleri ve haftadaki en geniş ders saati dizini listesi."""
    gunler = [
        g for g in db.scalars(
            select(Day).options(selectinload(Day.periods))
            .where(Day.term_id == donem.id, Day.is_active.is_(True))
            .order_by(Day.index)
        )
    ]
    en_fazla = max(
        (max((p.index for p in g.periods), default=-1) for g in gunler), default=-1
    )
    return gunler, list(range(en_fazla + 1))


def _tablolar(db: Session, timetable_id: int, bakis: str) -> dict[str, dict]:
    """Anahtar (şube adı ya da öğretmen adı) -> {(gun, ders): hücre}"""
    hucreler = izgara_hucreleri(db, timetable_id)
    gruplar: dict[str, dict] = defaultdict(dict)
    for h in hucreler:
        anahtar = h.section_name if bakis == "sube" else h.teacher_name
        gruplar[anahtar][(h.day_index, h.period_index)] = h
    return dict(sorted(gruplar.items()))


def _baslik(db: Session, timetable_id: int) -> tuple[Timetable, str]:
    t = db.get(Timetable, timetable_id)
    if t is None or t.is_deleted:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Ders programı bulunamadı.")
    kurum = db.scalar(select(Institution).limit(1))
    return t, (kurum.name if kurum else "")


def _html(db: Session, timetable_id: int, bakis: str, donem: Term) -> str:
    t, kurum_adi = _baslik(db, timetable_id)
    gunler, ders_indexleri = _izgara_yapisi(db, donem)
    gruplar = _tablolar(db, timetable_id, bakis)

    parcalar = [
        "<style>",
        "@page{size:A4 landscape;margin:12mm}",
        "html{color-scheme:light}",
        "body{font-family:'Helvetica Neue',Arial,sans-serif;font-size:11px;"
        "color:#0f172a;background:#fff;margin:0}",
        "h1{font-size:16px;margin:0 0 2px}h2{font-size:13px;margin:0 0 8px;color:#475569;font-weight:500}",
        "section{page-break-after:always}section:last-child{page-break-after:auto}",
        "table{border-collapse:collapse;width:100%}",
        "th,td{border:1px solid #cbd5e1;padding:5px 6px;text-align:center;vertical-align:middle;height:34px}",
        "th{background:#f1f5f9;font-weight:600}",
        "td .ders{font-weight:600}td .alt{font-size:9px;color:#64748b}",
        "</style>",
    ]
    for anahtar, hucre_map in gruplar.items():
        parcalar.append("<section>")
        parcalar.append(f"<h1>{_kacis(anahtar)}</h1>")
        parcalar.append(f"<h2>{_kacis(kurum_adi)} · {_kacis(t.name)}</h2>")
        parcalar.append("<table><thead><tr><th></th>")
        for g in gunler:
            parcalar.append(f"<th>{_kacis(g.name)}</th>")
        parcalar.append("</tr></thead><tbody>")
        for di in ders_indexleri:
            parcalar.append(f"<tr><th>{di + 1}. ders</th>")
            for g in gunler:
                h = hucre_map.get((g.index, di))
                if h is None:
                    parcalar.append("<td></td>")
                else:
                    alt = h.teacher_name if bakis == "sube" else h.section_name
                    parcalar.append(
                        f'<td style="background:{h.subject_color}22">'
                        f'<div class="ders">{_kacis(h.subject_name)}</div>'
                        f'<div class="alt">{_kacis(alt)}</div></td>'
                    )
            parcalar.append("</tr>")
        parcalar.append("</tbody></table></section>")
    if not gruplar:
        parcalar.append("<p>Bu programda yerleşmiş ders yok.</p>")
    return "".join(parcalar)


def _carsaf_html(db: Session, timetable_id: int, bakis: str, donem: Term) -> str:
    """Tüm şubeleri (ya da öğretmenleri) tek sayfada gösteren toplu liste.

    Satırlar şube/öğretmen, sütunlar gün × ders saati. Hücrelerde yer dar
    olduğu için tanımlıysa kısa kodlar kullanılır.
    """
    t, kurum_adi = _baslik(db, timetable_id)
    gunler, _ = _izgara_yapisi(db, donem)
    gruplar = _tablolar(db, timetable_id, bakis)

    # Her günün kendi ders saati dizini listesi — günler farklı uzunlukta olabilir.
    gun_saatleri = [
        (g, sorted(p.index for p in g.periods)) for g in gunler
    ]
    gun_saatleri = [(g, idx) for g, idx in gun_saatleri if idx]
    sutun_sayisi = sum(len(idx) for _, idx in gun_saatleri)

    # Sütun sayısı arttıkça yazı küçülür; A4 yatay sayfaya sığması için.
    punto = 7.5 if sutun_sayisi <= 25 else 6.5 if sutun_sayisi <= 35 else 5.5

    p: list[str] = [
        "<style>",
        "@page{size:A4 landscape;margin:8mm}",
        "html{color-scheme:light}",
        "body{font-family:'Helvetica Neue',Arial,sans-serif;color:#0f172a;"
        "background:#fff;margin:0}",
        "h1{font-size:13px;margin:0 0 1px}",
        "h2{font-size:10px;margin:0 0 6px;color:#475569;font-weight:500}",
        f"table{{border-collapse:collapse;width:100%;table-layout:fixed;font-size:{punto}px}}",
        "th,td{border:1px solid #cbd5e1;padding:1px;text-align:center;"
        "overflow:hidden;background:#fff}",
        "th{background:#f1f5f9;font-weight:600}",
        "th.ad{width:70px;text-align:left;padding-left:4px}",
        "td.ad{text-align:left;padding-left:4px;font-weight:600;background:#f8fafc}",
        "td.tnf{background:#e2e8f0}",
        "th.gun{border-left:2px solid #64748b}",
        "td.gunbas,th.gunbas{border-left:2px solid #64748b}",
        ".ders{font-weight:600;display:block;line-height:1.15}",
        ".alt{color:#475569;display:block;line-height:1.15}",
        "</style>",
        f"<h1>{_kacis(t.name)} — Çarşaf Liste "
        f"({'Şube' if bakis == 'sube' else 'Öğretmen'})</h1>",
        f"<h2>{_kacis(kurum_adi)}</h2>",
        "<table><thead><tr>",
        f'<th class="ad" rowspan="2">{"Şube" if bakis == "sube" else "Öğretmen"}</th>',
    ]
    for g, idx in gun_saatleri:
        p.append(f'<th class="gun" colspan="{len(idx)}">{_kacis(g.name)}</th>')
    p.append("</tr><tr>")
    for g, idx in gun_saatleri:
        for konum, di in enumerate(idx):
            sinif = ' class="gunbas"' if konum == 0 else ""
            p.append(f"<th{sinif}>{di + 1}</th>")
    p.append("</tr></thead><tbody>")

    for anahtar, hucre_map in gruplar.items():
        p.append(f'<tr><td class="ad">{_kacis(anahtar)}</td>')
        for g, idx in gun_saatleri:
            for konum, di in enumerate(idx):
                sinif = "gunbas" if konum == 0 else ""
                period = next((x for x in g.periods if x.index == di), None)
                if period is not None and period.is_break:
                    p.append(f'<td class="tnf {sinif}"></td>')
                    continue
                h = hucre_map.get((g.index, di))
                if h is None:
                    p.append(f'<td class="{sinif}"></td>')
                    continue
                ders = h.subject_short or h.subject_name
                alt = (
                    (h.teacher_short or h.teacher_name)
                    if bakis == "sube"
                    else h.section_name
                )
                p.append(
                    f'<td class="{sinif}" style="background:{h.subject_color}1f">'
                    f'<span class="ders">{_kacis(ders)}</span>'
                    f'<span class="alt">{_kacis(alt)}</span></td>'
                )
        p.append("</tr>")

    p.append("</tbody></table>")
    if not gruplar:
        p.append("<p>Bu programda yerleşmiş ders yok.</p>")
    return "".join(p)


def _kacis(s: str) -> str:
    return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def _icerik(db: Session, timetable_id: int, bakis: str, duzen: str, donem: Term) -> str:
    if duzen == "carsaf":
        return _carsaf_html(db, timetable_id, bakis, donem)
    return _html(db, timetable_id, bakis, donem)


@router.get("/html", response_class=Response)
def html_cikti(
    timetable_id: int,
    bakis: str = Query("sube", pattern="^(sube|ogretmen)$"),
    duzen: str = Query("ayri", pattern="^(ayri|carsaf)$"),
    db: Session = Depends(get_db),
    donem: Term = Depends(aktif_donem),
) -> Response:
    return Response(
        _icerik(db, timetable_id, bakis, duzen, donem), media_type="text/html; charset=utf-8"
    )


@router.get("/pdf", response_class=Response)
def pdf_cikti(
    timetable_id: int,
    bakis: str = Query("sube", pattern="^(sube|ogretmen)$"),
    duzen: str = Query("ayri", pattern="^(ayri|carsaf)$"),
    db: Session = Depends(get_db),
    donem: Term = Depends(aktif_donem),
) -> Response:
    try:
        from weasyprint import HTML
    except (ImportError, OSError) as e:
        raise HTTPException(
            status.HTTP_503_SERVICE_UNAVAILABLE,
            "PDF üretimi için gereken sistem kütüphaneleri kurulu değil "
            f"(macOS: brew install pango). Ayrıntı: {e}. "
            "Bu arada HTML çıktısını tarayıcıdan yazdırabilirsiniz.",
        )
    pdf = HTML(string=_icerik(db, timetable_id, bakis, duzen, donem)).write_pdf()
    ad = f"ders-programi-{duzen}-{bakis}.pdf"
    return Response(pdf, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{ad}"'})


@router.get("/xlsx", response_class=Response)
def excel_cikti(
    timetable_id: int,
    bakis: str = Query("sube", pattern="^(sube|ogretmen)$"),
    duzen: str = Query("ayri", pattern="^(ayri|carsaf)$"),
    db: Session = Depends(get_db),
    donem: Term = Depends(aktif_donem),
) -> Response:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side

    gunler, ders_indexleri = _izgara_yapisi(db, donem)
    gruplar = _tablolar(db, timetable_id, bakis)

    wb = Workbook()
    wb.remove(wb.active)
    kenar = Border(*[Side(style="thin", color="CBD5E1")] * 4)
    ortala = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if duzen == "carsaf":
        _carsaf_excel(wb, gunler, gruplar, bakis, kenar, ortala, Font, Alignment)
        return _excel_yanit(wb, f"carsaf-{bakis}")

    for anahtar, hucre_map in gruplar.items():
        ws = wb.create_sheet(title=anahtar[:31].replace("/", "-"))
        ws.cell(row=1, column=1, value="").border = kenar
        for c, g in enumerate(gunler, start=2):
            h = ws.cell(row=1, column=c, value=g.name)
            h.font, h.alignment, h.border = Font(bold=True), ortala, kenar
            ws.column_dimensions[h.column_letter].width = 24
        for r, di in enumerate(ders_indexleri, start=2):
            b = ws.cell(row=r, column=1, value=f"{di + 1}. ders")
            b.font, b.alignment, b.border = Font(bold=True), ortala, kenar
            ws.row_dimensions[r].height = 32
            for c, g in enumerate(gunler, start=2):
                hucre = hucre_map.get((g.index, di))
                metin = ""
                if hucre is not None:
                    alt = hucre.teacher_name if bakis == "sube" else hucre.section_name
                    metin = f"{hucre.subject_name}\n{alt}"
                x = ws.cell(row=r, column=c, value=metin)
                x.alignment, x.border = ortala, kenar

    if not gruplar:
        wb.create_sheet(title="Bos")["A1"] = "Bu programda yerleşmiş ders yok."

    return _excel_yanit(wb, f"ders-programi-{bakis}")


def _carsaf_excel(wb, gunler, gruplar, bakis, kenar, ortala, Font, Alignment) -> None:
    """Tek sayfada toplu liste: satırlar şube/öğretmen, sütunlar gün × ders saati."""
    ws = wb.create_sheet(title="Çarşaf")
    gun_saatleri = [(g, sorted(p.index for p in g.periods)) for g in gunler]
    gun_saatleri = [(g, idx) for g, idx in gun_saatleri if idx]

    ws.column_dimensions["A"].width = 18
    ws.freeze_panes = "B3"

    baslik = ws.cell(row=1, column=1, value="Şube" if bakis == "sube" else "Öğretmen")
    baslik.font, baslik.border = Font(bold=True), kenar
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    baslik.alignment = Alignment(horizontal="left", vertical="center")

    sutun = 2
    for g, idx in gun_saatleri:
        gun_hucre = ws.cell(row=1, column=sutun, value=g.name)
        gun_hucre.font, gun_hucre.alignment, gun_hucre.border = (
            Font(bold=True), ortala, kenar,
        )
        if len(idx) > 1:
            ws.merge_cells(start_row=1, start_column=sutun,
                           end_row=1, end_column=sutun + len(idx) - 1)
        for konum, di in enumerate(idx):
            h = ws.cell(row=2, column=sutun + konum, value=di + 1)
            h.font, h.alignment, h.border = Font(bold=True), ortala, kenar
            ws.column_dimensions[h.column_letter].width = 12
        sutun += len(idx)

    for satir, (anahtar, hucre_map) in enumerate(gruplar.items(), start=3):
        ad = ws.cell(row=satir, column=1, value=anahtar)
        ad.font, ad.border = Font(bold=True), kenar
        ws.row_dimensions[satir].height = 30
        sutun = 2
        for g, idx in gun_saatleri:
            for konum, di in enumerate(idx):
                period = next((x for x in g.periods if x.index == di), None)
                h = hucre_map.get((g.index, di))
                metin = ""
                if period is not None and period.is_break:
                    metin = "—"
                elif h is not None:
                    alt = h.teacher_name if bakis == "sube" else h.section_name
                    metin = f"{h.subject_name}\n{alt}"
                x = ws.cell(row=satir, column=sutun + konum, value=metin)
                x.alignment, x.border = ortala, kenar
            sutun += len(idx)

    if not gruplar:
        ws["A3"] = "Bu programda yerleşmiş ders yok."


def _excel_yanit(wb, dosya_adi: str) -> Response:
    tampon = io.BytesIO()
    wb.save(tampon)
    return Response(
        tampon.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{dosya_adi}.xlsx"'},
    )
