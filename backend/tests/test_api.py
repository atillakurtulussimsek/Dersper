"""Uçtan uca API testleri: kurulumdan yayına kadar tüm akış."""
from fastapi.testclient import TestClient

from tests.conftest import cozumsuz_calistir, uret_ve_bekle


def _tanimlar(c: TestClient, ek: str = "") -> dict:
    """Küçük bir okul kurar ve kimlikleri döner. `ek`, adları benzersizleştirir."""
    gunler = [g for g in c.get("/api/timegrid").json() if g["is_active"]]
    dersler = [
        c.post("/api/subjects", json={"name": ad}).json()["id"]
        for ad in (f"Matematik{ek}", f"Türkçe{ek}", f"Fen{ek}", f"Sosyal{ek}")
    ]
    ogretmenler = [
        c.post("/api/teachers", json={"full_name": ad}).json()["id"]
        for ad in (f"Ayşe{ek}", f"Mehmet{ek}", f"Zeynep{ek}", f"Ali{ek}")
    ]
    subeler = [
        c.post("/api/sections", json={"name": ad, "grade_level": 5}).json()["id"]
        for ad in (f"5-A{ek}", f"5-B{ek}")
    ]
    for sube in subeler:
        for i, ders in enumerate(dersler):
            c.post("/api/curriculum", json={
                "section_id": sube, "subject_id": ders, "teacher_id": ogretmenler[i],
                "weekly_hours": 5 if i < 2 else 4,
                "block_pattern": "2+2" if i == 2 else "", "max_per_day": 2,
            })
    return {"gunler": gunler, "dersler": dersler, "ogretmenler": ogretmenler,
            "subeler": subeler}


def test_kurulum_ve_oturum(yonetici: TestClient):
    assert yonetici.get("/api/auth/status").json()["has_institutions"] is True
    # kurulumda verilen parolayla yeniden giriş yapılabilir
    r = yonetici.post("/api/auth/login", json={
        "email": "yonetici@ornek.com", "password": "parola1234",
    })
    assert r.status_code == 200
    assert yonetici.get("/api/auth/me").json()["email"] == "yonetici@ornek.com"
    assert yonetici.get("/api/institution").json()["name"] == "Test Ortaokulu"


def test_ayni_eposta_ikinci_kez_kayit_olamaz(yonetici: TestClient):
    r = yonetici.post("/api/auth/register", json={
        "institution_name": "İkinci Okul", "institution_type": "k12",
        "full_name": "Biri", "email": "yonetici@ornek.com", "password": "parola1234",
    })
    assert r.status_code == 409
    assert "yalnızca bir kuruma" in r.text


def test_jetonsuz_erisim_engellenir(istemci: TestClient):
    istemci.headers.pop("Authorization", None)
    assert istemci.get("/api/teachers").status_code == 401


def test_varsayilan_zaman_izgarasi(yonetici: TestClient):
    aktif = [g for g in yonetici.get("/api/timegrid").json() if g["is_active"]]
    assert len(aktif) == 5                      # Pazartesi–Cuma
    assert len(aktif[0]["periods"]) == 8


def test_musaitlik_kaydedilir(yonetici: TestClient):
    ogretmen = yonetici.post("/api/teachers", json={"full_name": "Müsaitlik Testi"}).json()
    cuma = [g for g in yonetici.get("/api/timegrid").json() if g["is_active"]][4]
    yonetici.put(f"/api/teachers/{ogretmen['id']}/availability", json={
        "cells": [{"period_id": p["id"], "state": "uygun_degil"} for p in cuma["periods"]]
    })
    kayitli = yonetici.get(f"/api/teachers/{ogretmen['id']}/availability").json()
    assert len(kayitli) == len(cuma["periods"])
    assert all(h["state"] == "uygun_degil" for h in kayitli)


def test_ayni_ders_ayni_ogretmenle_iki_kez_atanamaz(yonetici: TestClient):
    d = yonetici.post("/api/subjects", json={"name": "Tekrar Dersi"}).json()["id"]
    o = yonetici.post("/api/teachers", json={"full_name": "Tekrar Öğretmeni"}).json()["id"]
    s = yonetici.post("/api/sections", json={"name": "9-Z"}).json()["id"]
    govde = {"section_id": s, "subject_id": d, "teacher_id": o, "weekly_hours": 2}
    assert yonetici.post("/api/curriculum", json=govde).status_code == 201
    r = yonetici.post("/api/curriculum", json=govde)
    assert r.status_code == 409
    assert "başka bir öğretmenle" in r.text


def test_ayni_ders_farkli_ogretmenle_atanabilir(yonetici: TestClient):
    """İngilizce'nin 2 saati bir, 2 saati başka öğretmende olabilir."""
    d = yonetici.post("/api/subjects", json={"name": "İngilizce"}).json()["id"]
    o1 = yonetici.post("/api/teachers", json={"full_name": "Ayşe Yılmaz"}).json()["id"]
    o2 = yonetici.post("/api/teachers", json={"full_name": "Mehmet Kaya"}).json()["id"]
    s = yonetici.post("/api/sections", json={"name": "9-A"}).json()["id"]

    for o in (o1, o2):
        r = yonetici.post("/api/curriculum", json={
            "section_id": s, "subject_id": d, "teacher_id": o, "weekly_hours": 2,
        })
        assert r.status_code == 201, r.text

    atamalar = yonetici.get(f"/api/curriculum?section_id={s}").json()
    assert len(atamalar) == 2
    assert {a["teacher"]["full_name"] for a in atamalar} == {"Ayşe Yılmaz", "Mehmet Kaya"}


def test_ayni_dersin_iki_ogretmeni_ayni_saate_yerlesmez(yonetici: TestClient):
    """İki ayrı atama olsa da şube aynı anda tek derste olur."""
    d = yonetici.post("/api/subjects", json={"name": "İngilizce"}).json()["id"]
    o1 = yonetici.post("/api/teachers", json={"full_name": "Ayşe Yılmaz"}).json()["id"]
    o2 = yonetici.post("/api/teachers", json={"full_name": "Mehmet Kaya"}).json()["id"]
    s = yonetici.post("/api/sections", json={"name": "9-A"}).json()["id"]
    for o in (o1, o2):
        yonetici.post("/api/curriculum", json={
            "section_id": s, "subject_id": d, "teacher_id": o, "weekly_hours": 3,
        })

    pid = yonetici.post("/api/timetables", json={"name": "Bölünmüş ders"}).json()["id"]
    deneme = uret_ve_bekle(yonetici, pid)
    assert deneme["status"] == "basarili", deneme["report"]

    hucreler = yonetici.get(f"/api/timetables/{pid}/grid").json()["cells"]
    assert len(hucreler) == 6
    saatler = [h["period_id"] for h in hucreler]
    assert len(saatler) == len(set(saatler))          # çakışma yok
    assert {h["teacher_name"] for h in hucreler} == {"Ayşe Yılmaz", "Mehmet Kaya"}


def test_kopyalama_ayni_dersi_farkli_ogretmenle_getirir(yonetici: TestClient):
    d = yonetici.post("/api/subjects", json={"name": "İngilizce"}).json()["id"]
    o1 = yonetici.post("/api/teachers", json={"full_name": "Ayşe Yılmaz"}).json()["id"]
    o2 = yonetici.post("/api/teachers", json={"full_name": "Mehmet Kaya"}).json()["id"]
    kaynak = yonetici.post("/api/sections", json={"name": "9-A"}).json()["id"]
    hedef = yonetici.post("/api/sections", json={"name": "9-B"}).json()["id"]

    e1 = yonetici.post("/api/curriculum", json={
        "section_id": kaynak, "subject_id": d, "teacher_id": o1, "weekly_hours": 2,
    }).json()["id"]
    # Hedefte aynı ders başka öğretmenle zaten var: kopyalama yine de eklenmeli.
    yonetici.post("/api/curriculum", json={
        "section_id": hedef, "subject_id": d, "teacher_id": o2, "weekly_hours": 2,
    })

    r = yonetici.post("/api/curriculum/copy", json={
        "entry_ids": [e1], "section_ids": [hedef],
    }).json()
    assert len(r["created"]) == 1
    assert len(yonetici.get(f"/api/curriculum?section_id={hedef}").json()) == 2

    # İkinci kez kopyalamak artık aynı üçlü olduğu için atlanır.
    tekrar = yonetici.post("/api/curriculum/copy", json={
        "entry_ids": [e1], "section_ids": [hedef],
    }).json()
    assert tekrar["created"] == []
    assert "Ayşe Yılmaz ile zaten tanımlı" in tekrar["skipped"][0]


def test_program_uretilir_ve_yayinlanir(yonetici: TestClient):
    _tanimlar(yonetici)
    pid = yonetici.post("/api/timetables", json={"name": "2026 Güz"}).json()["id"]

    deneme = uret_ve_bekle(yonetici, pid)
    assert deneme["status"] == "basarili", deneme["report"]

    izgara = yonetici.get(f"/api/timetables/{pid}/grid").json()
    assert len(izgara["cells"]) == 36           # 2 şube × 18 saat

    # Dolu bir saate taşımak reddedilir.
    h = izgara["cells"][0]
    diger = next(x for x in izgara["cells"]
                 if x["section_id"] == h["section_id"] and x["period_id"] != h["period_id"])
    r = yonetici.patch(f"/api/timetables/{pid}/assignments/{h['assignment_id']}",
                       json={"period_id": diger["period_id"]})
    assert r.status_code == 409

    yayin = yonetici.post(f"/api/timetables/{pid}/publish").json()
    assert yayin["public_token"]

    # Yayınlanan program girişsiz okunabilir.
    yonetici.headers.pop("Authorization")
    acik = yonetici.get(f"/api/public/timetables/{yayin['public_token']}").json()
    assert len(acik["cells"]) == 36


def test_cikti_bicimleri(yonetici: TestClient):
    _tanimlar(yonetici)
    pid = yonetici.post("/api/timetables", json={"name": "Çıktı"}).json()["id"]
    uret_ve_bekle(yonetici, pid)
    assert yonetici.get(f"/api/timetables/{pid}/export/xlsx?bakis=sube").status_code == 200
    assert yonetici.get(f"/api/timetables/{pid}/export/html?bakis=ogretmen").status_code == 200


def test_cozumsuz_program_tani_raporu_uretir(yonetici: TestClient):
    """Tek öğretmen iki şubede haftalık 50 saat: ızgarada 40 saat var, sığmaz."""
    d = yonetici.post("/api/subjects", json={"name": "Aşırı Yük"}).json()["id"]
    o = yonetici.post("/api/teachers", json={"full_name": "Yorgun Öğretmen"}).json()["id"]
    for ad in ("12-Y", "12-Z"):
        s = yonetici.post("/api/sections", json={"name": ad}).json()["id"]
        r = yonetici.post("/api/curriculum", json={
            "section_id": s, "subject_id": d, "teacher_id": o,
            "weekly_hours": 25, "max_per_day": 8,
        })
        assert r.status_code == 201, r.text
    pid = yonetici.post("/api/timetables", json={"name": "Çözümsüz"}).json()["id"]
    deneme = cozumsuz_calistir(yonetici, pid)

    # Çözümsüz iş kendiliğinden durmaz; durdurulana kadar denemeye devam eder.
    assert deneme["status"] == "durduruldu"
    assert deneme["attempts"] >= 1
    kodlar = {b["kod"] for b in deneme["report"]["bulgular"]}
    assert "ogretmen_kapasite" in kodlar
    assert deneme["report"]["ozet"]["yerlesmeyen_toplam"] > 0
    # Yapay zeka kapalıyken açıklama üretilmez ama çalıştırma yine de kapanır.
    assert deneme["ai_explanation"] is None


def test_yapay_zeka_anahtari_geri_okunmaz(yonetici: TestClient):
    assert yonetici.get("/api/ai/settings").json()["has_api_key"] is False
    yonetici.put("/api/ai/settings", json={
        "enabled": True, "base_url": "http://yok.local/v1",
        "model": "test", "api_key": "sk-gizli-anahtar-1234",
    })
    ayar = yonetici.get("/api/ai/settings").json()
    assert ayar["has_api_key"] is True
    assert ayar["api_key_masked"] == "sk-gi…1234"
    assert "api_key" not in ayar          # ham anahtar hiçbir zaman dönmez


def test_sube_musaitligi_kaydedilir(yonetici: TestClient):
    sube = yonetici.post("/api/sections", json={"name": "6-A"}).json()
    gunler = [g for g in yonetici.get("/api/timegrid").json() if g["is_active"]]
    sabah = [p for g in gunler for p in g["periods"] if p["index"] < 4]

    yonetici.put(f"/api/sections/{sube['id']}/availability", json={
        "cells": [{"period_id": p["id"], "state": "uygun_degil"} for p in sabah]
    })
    kayitli = yonetici.get(f"/api/sections/{sube['id']}/availability").json()
    assert len(kayitli) == len(sabah)
    assert all(h["state"] == "uygun_degil" for h in kayitli)


def test_aksamci_sube_sabaha_ders_almaz(yonetici: TestClient):
    """Sabah saatleri kapatılan şubenin dersleri yalnızca öğleden sonraya yerleşir."""
    d = yonetici.post("/api/subjects", json={"name": "Matematik"}).json()["id"]
    o = yonetici.post("/api/teachers", json={"full_name": "Akşam Öğretmeni"}).json()["id"]
    s = yonetici.post("/api/sections", json={"name": "9-Akşam"}).json()["id"]

    gunler = [g for g in yonetici.get("/api/timegrid").json() if g["is_active"]]
    sabah = {p["id"] for g in gunler for p in g["periods"] if p["index"] < 4}
    yonetici.put(f"/api/sections/{s}/availability", json={
        "cells": [{"period_id": pid, "state": "uygun_degil"} for pid in sabah]
    })

    yonetici.post("/api/curriculum", json={
        "section_id": s, "subject_id": d, "teacher_id": o,
        "weekly_hours": 10, "max_per_day": 2,
    })
    pid = yonetici.post("/api/timetables", json={"name": "Akşam"}).json()["id"]
    deneme = uret_ve_bekle(yonetici, pid)
    assert deneme["status"] == "basarili", deneme["report"]

    hucreler = yonetici.get(f"/api/timetables/{pid}/grid").json()["cells"]
    assert len(hucreler) == 10
    assert all(h["period_index"] >= 4 for h in hucreler)


def test_asiri_kapali_sube_tani_raporunda_gorunur(yonetici: TestClient):
    d = yonetici.post("/api/subjects", json={"name": "Türkçe"}).json()["id"]
    o = yonetici.post("/api/teachers", json={"full_name": "Bir Öğretmen"}).json()["id"]
    s = yonetici.post("/api/sections", json={"name": "10-Dar"}).json()["id"]

    gunler = [g for g in yonetici.get("/api/timegrid").json() if g["is_active"]]
    # Şubeye haftada yalnızca 5 saat bırak.
    acik = {g["periods"][0]["id"] for g in gunler}
    kapali = [p["id"] for g in gunler for p in g["periods"] if p["id"] not in acik]
    yonetici.put(f"/api/sections/{s}/availability", json={
        "cells": [{"period_id": pid, "state": "uygun_degil"} for pid in kapali]
    })

    yonetici.post("/api/curriculum", json={
        "section_id": s, "subject_id": d, "teacher_id": o,
        "weekly_hours": 12, "max_per_day": 3,
    })
    pid = yonetici.post("/api/timetables", json={"name": "Dar"}).json()["id"]
    deneme = cozumsuz_calistir(yonetici, pid)

    assert deneme["status"] == "durduruldu"
    bulgu = next(b for b in deneme["report"]["bulgular"] if b["kod"] == "sube_kapasite")
    assert bulgu["mevcut"] == 5
    assert bulgu["gereken"] == 12


def test_blok_deseni_kaydedilir_ve_dogrulanir(yonetici: TestClient):
    d = yonetici.post("/api/subjects", json={"name": "Matematik"}).json()["id"]
    o = yonetici.post("/api/teachers", json={"full_name": "Desen Öğretmeni"}).json()["id"]
    s = yonetici.post("/api/sections", json={"name": "7-D"}).json()["id"]
    temel = {"section_id": s, "subject_id": d, "teacher_id": o, "weekly_hours": 5}

    # Toplamı tutmayan desen reddedilir.
    hatali = yonetici.post("/api/curriculum", json={**temel, "block_pattern": "2+2"})
    assert hatali.status_code == 422
    assert "eşit olmalı" in hatali.text

    # Geçerli desen tek biçime getirilerek saklanır.
    ok = yonetici.post("/api/curriculum", json={**temel, "block_pattern": "2, 2 ,1"})
    assert ok.status_code == 201
    assert ok.json()["block_pattern"] == "2+2+1"


def test_desen_bos_birakilirsa_tek_saatlere_acilir(yonetici: TestClient):
    d = yonetici.post("/api/subjects", json={"name": "Müzik"}).json()["id"]
    o = yonetici.post("/api/teachers", json={"full_name": "Müzik Öğretmeni"}).json()["id"]
    s = yonetici.post("/api/sections", json={"name": "7-E"}).json()["id"]
    r = yonetici.post("/api/curriculum", json={
        "section_id": s, "subject_id": d, "teacher_id": o, "weekly_hours": 3,
    })
    assert r.json()["block_pattern"] == "1+1+1"


def test_istenen_desen_programa_yansir(yonetici: TestClient):
    """3+2 istenen 5 saatlik ders, bir gün 3 diğer gün 2 saat olarak yerleşir."""
    d = yonetici.post("/api/subjects", json={"name": "Atölye"}).json()["id"]
    o = yonetici.post("/api/teachers", json={"full_name": "Atölye Öğretmeni"}).json()["id"]
    s = yonetici.post("/api/sections", json={"name": "11-A"}).json()["id"]
    yonetici.post("/api/curriculum", json={
        "section_id": s, "subject_id": d, "teacher_id": o,
        "weekly_hours": 5, "block_pattern": "3+2", "max_per_day": 3,
    })
    pid = yonetici.post("/api/timetables", json={"name": "Desen"}).json()["id"]
    deneme = uret_ve_bekle(yonetici, pid)
    assert deneme["status"] == "basarili", deneme["report"]

    hucreler = yonetici.get(f"/api/timetables/{pid}/grid").json()["cells"]
    gunluk: dict[int, int] = {}
    for h in hucreler:
        gunluk[h["day_index"]] = gunluk.get(h["day_index"], 0) + 1
    assert sorted(gunluk.values(), reverse=True) == [3, 2]


def _kopyalama_ortami(c: TestClient) -> dict:
    d1 = c.post("/api/subjects", json={"name": "Matematik"}).json()["id"]
    d2 = c.post("/api/subjects", json={"name": "Türkçe"}).json()["id"]
    o = c.post("/api/teachers", json={"full_name": "Ayşe Yılmaz"}).json()["id"]
    kaynak = c.post("/api/sections", json={"name": "5-A"}).json()["id"]
    hedef1 = c.post("/api/sections", json={"name": "5-B"}).json()["id"]
    hedef2 = c.post("/api/sections", json={"name": "5-C"}).json()["id"]

    e1 = c.post("/api/curriculum", json={
        "section_id": kaynak, "subject_id": d1, "teacher_id": o,
        "weekly_hours": 5, "block_pattern": "2+2+1", "max_per_day": 2,
    }).json()["id"]
    e2 = c.post("/api/curriculum", json={
        "section_id": kaynak, "subject_id": d2, "teacher_id": o,
        "weekly_hours": 4, "max_per_day": 1,
    }).json()["id"]
    return {"d1": d1, "d2": d2, "o": o, "kaynak": kaynak,
            "hedef1": hedef1, "hedef2": hedef2, "e1": e1, "e2": e2}


def test_tek_ders_baska_subelere_kopyalanir(yonetici: TestClient):
    v = _kopyalama_ortami(yonetici)
    r = yonetici.post("/api/curriculum/copy", json={
        "entry_ids": [v["e1"]], "section_ids": [v["hedef1"], v["hedef2"]],
    })
    assert r.status_code == 201
    olusan = r.json()["created"]
    assert len(olusan) == 2
    assert {k["section_id"] for k in olusan} == {v["hedef1"], v["hedef2"]}

    # Şube dışındaki her şey aynı kalır.
    for k in olusan:
        assert k["subject_id"] == v["d1"]
        assert k["teacher_id"] == v["o"]
        assert k["weekly_hours"] == 5
        assert k["block_pattern"] == "2+2+1"
        assert k["max_per_day"] == 2


def test_tum_mufredat_kopyalanir(yonetici: TestClient):
    v = _kopyalama_ortami(yonetici)
    r = yonetici.post("/api/curriculum/copy", json={
        "entry_ids": [v["e1"], v["e2"]], "section_ids": [v["hedef1"]],
    }).json()
    assert len(r["created"]) == 2
    hedefin = yonetici.get(f"/api/curriculum?section_id={v['hedef1']}").json()
    assert {m["subject"]["name"] for m in hedefin} == {"Matematik", "Türkçe"}


def test_zaten_tanimli_ders_atlanir(yonetici: TestClient):
    v = _kopyalama_ortami(yonetici)
    yonetici.post("/api/curriculum", json={
        "section_id": v["hedef1"], "subject_id": v["d1"], "teacher_id": v["o"],
        "weekly_hours": 2,
    })
    r = yonetici.post("/api/curriculum/copy", json={
        "entry_ids": [v["e1"]], "section_ids": [v["hedef1"], v["hedef2"]],
    }).json()

    assert len(r["created"]) == 1                 # yalnızca 5-C'ye kopyalandı
    assert len(r["skipped"]) == 1
    assert "zaten tanımlı" in r["skipped"][0]
    # Var olan satır değişmeden kalır.
    kalan = yonetici.get(f"/api/curriculum?section_id={v['hedef1']}").json()
    assert kalan[0]["weekly_hours"] == 2


def test_kaynak_subeye_kopyalama_atlanir(yonetici: TestClient):
    v = _kopyalama_ortami(yonetici)
    r = yonetici.post("/api/curriculum/copy", json={
        "entry_ids": [v["e1"]], "section_ids": [v["kaynak"]],
    }).json()
    assert r["created"] == []
    assert "kaynak şubenin kendisi" in r["skipped"][0]


def test_kopyalanan_satir_bagimsiz_duzenlenir(yonetici: TestClient):
    v = _kopyalama_ortami(yonetici)
    kopya = yonetici.post("/api/curriculum/copy", json={
        "entry_ids": [v["e1"]], "section_ids": [v["hedef1"]],
    }).json()["created"][0]

    yonetici.put(f"/api/curriculum/{kopya['id']}", json={
        "section_id": kopya["section_id"], "subject_id": v["d1"], "teacher_id": v["o"],
        "weekly_hours": 3, "block_pattern": "2+1", "max_per_day": 2,
    })
    kaynak = yonetici.get(f"/api/curriculum?section_id={v['kaynak']}").json()
    asil = next(m for m in kaynak if m["id"] == v["e1"])
    assert asil["weekly_hours"] == 5              # kaynak etkilenmez
    assert asil["block_pattern"] == "2+2+1"


def test_bilinmeyen_satir_kopyalanamaz(yonetici: TestClient):
    v = _kopyalama_ortami(yonetici)
    assert yonetici.post("/api/curriculum/copy", json={
        "entry_ids": [9999], "section_ids": [v["hedef1"]],
    }).status_code == 404
    assert yonetici.post("/api/curriculum/copy", json={
        "entry_ids": [v["e1"]], "section_ids": [9999],
    }).status_code == 404


def _saatler(c: TestClient) -> list[dict]:
    return [p for g in c.get("/api/timegrid").json() if g["is_active"] for p in g["periods"]]


def test_sube_musaitligi_baska_subelere_kopyalanir(yonetici: TestClient):
    kaynak = yonetici.post("/api/sections", json={"name": "5-A"}).json()["id"]
    h1 = yonetici.post("/api/sections", json={"name": "5-B"}).json()["id"]
    h2 = yonetici.post("/api/sections", json={"name": "5-C"}).json()["id"]

    sabah = [p for p in _saatler(yonetici) if p["index"] < 4]
    yonetici.put(f"/api/sections/{kaynak}/availability", json={
        "cells": [{"period_id": p["id"], "state": "uygun_degil"} for p in sabah]
    })

    r = yonetici.post(f"/api/sections/{kaynak}/availability/copy",
                      json={"section_ids": [h1, h2]}).json()
    assert r["copied_to"] == ["5-B", "5-C"]
    assert r["cells"] == len(sabah)

    for hedef in (h1, h2):
        kopya = yonetici.get(f"/api/sections/{hedef}/availability").json()
        assert {h["period_id"] for h in kopya} == {p["id"] for p in sabah}
        assert all(h["state"] == "uygun_degil" for h in kopya)


def test_kopyalama_hedefin_onceki_planini_siler(yonetici: TestClient):
    """Birleştirme yok: hedefin eski işaretlemesi tamamen gider."""
    kaynak = yonetici.post("/api/sections", json={"name": "6-A"}).json()["id"]
    hedef = yonetici.post("/api/sections", json={"name": "6-B"}).json()["id"]
    saatler = _saatler(yonetici)

    # Hedefte tamamen farklı bir plan var.
    aksam = [p for p in saatler if p["index"] >= 4]
    yonetici.put(f"/api/sections/{hedef}/availability", json={
        "cells": [{"period_id": p["id"], "state": "uygun_degil"} for p in aksam]
    })
    # Kaynakta sabah kapalı.
    sabah = [p for p in saatler if p["index"] < 4]
    yonetici.put(f"/api/sections/{kaynak}/availability", json={
        "cells": [{"period_id": p["id"], "state": "uygun_degil"} for p in sabah]
    })

    yonetici.post(f"/api/sections/{kaynak}/availability/copy",
                  json={"section_ids": [hedef]})

    sonuc = yonetici.get(f"/api/sections/{hedef}/availability").json()
    assert {h["period_id"] for h in sonuc} == {p["id"] for p in sabah}
    assert not ({h["period_id"] for h in sonuc} & {p["id"] for p in aksam})


def test_bos_musaitlik_kopyalanirsa_hedef_temizlenir(yonetici: TestClient):
    kaynak = yonetici.post("/api/sections", json={"name": "7-A"}).json()["id"]
    hedef = yonetici.post("/api/sections", json={"name": "7-B"}).json()["id"]

    yonetici.put(f"/api/sections/{hedef}/availability", json={
        "cells": [{"period_id": p["id"], "state": "uygun_degil"} for p in _saatler(yonetici)]
    })
    r = yonetici.post(f"/api/sections/{kaynak}/availability/copy",
                      json={"section_ids": [hedef]}).json()
    assert r["cells"] == 0
    assert yonetici.get(f"/api/sections/{hedef}/availability").json() == []


def test_kaynak_sube_hedeflerden_cikarilir(yonetici: TestClient):
    kaynak = yonetici.post("/api/sections", json={"name": "8-A"}).json()["id"]
    hedef = yonetici.post("/api/sections", json={"name": "8-B"}).json()["id"]

    r = yonetici.post(f"/api/sections/{kaynak}/availability/copy",
                      json={"section_ids": [kaynak, hedef]}).json()
    assert r["copied_to"] == ["8-B"]

    # Yalnızca kendisi hedefse istek reddedilir.
    assert yonetici.post(f"/api/sections/{kaynak}/availability/copy",
                         json={"section_ids": [kaynak]}).status_code == 404


def test_tercih_durumu_da_kopyalanir(yonetici: TestClient):
    kaynak = yonetici.post("/api/sections", json={"name": "9-A"}).json()["id"]
    hedef = yonetici.post("/api/sections", json={"name": "9-B"}).json()["id"]
    saatler = _saatler(yonetici)

    yonetici.put(f"/api/sections/{kaynak}/availability", json={"cells": [
        {"period_id": saatler[0]["id"], "state": "uygun_degil"},
        {"period_id": saatler[1]["id"], "state": "tercih"},
    ]})
    yonetici.post(f"/api/sections/{kaynak}/availability/copy",
                  json={"section_ids": [hedef]})

    kopya = {h["period_id"]: h["state"] for h in
             yonetici.get(f"/api/sections/{hedef}/availability").json()}
    assert kopya == {saatler[0]["id"]: "uygun_degil", saatler[1]["id"]: "tercih"}


def test_carsaf_html_tum_subeleri_tek_tabloda_verir(yonetici: TestClient):
    _tanimlar(yonetici)
    pid = yonetici.post("/api/timetables", json={"name": "Çarşaf"}).json()["id"]
    uret_ve_bekle(yonetici, pid)

    r = yonetici.get(f"/api/timetables/{pid}/export/html?bakis=sube&duzen=carsaf")
    assert r.status_code == 200
    govde = r.text
    assert "Çarşaf Liste (Şube)" in govde
    assert govde.count("<table") == 1          # ayrı sayfalar değil, tek tablo
    for sube in ("5-A", "5-B"):
        assert sube in govde
    assert "Pazartesi" in govde and "Cuma" in govde


def test_carsaf_ogretmen_bakisi(yonetici: TestClient):
    _tanimlar(yonetici)
    pid = yonetici.post("/api/timetables", json={"name": "Çarşaf"}).json()["id"]
    uret_ve_bekle(yonetici, pid)

    govde = yonetici.get(
        f"/api/timetables/{pid}/export/html?bakis=ogretmen&duzen=carsaf"
    ).text
    assert "Çarşaf Liste (Öğretmen)" in govde
    assert "Ayşe" in govde


def test_carsaf_kisa_kod_kullanir(yonetici: TestClient):
    d = yonetici.post("/api/subjects",
                      json={"name": "Matematik", "short_code": "MAT"}).json()["id"]
    o = yonetici.post("/api/teachers",
                      json={"full_name": "Ayşe Yılmaz", "short_code": "AY"}).json()["id"]
    s = yonetici.post("/api/sections", json={"name": "5-A"}).json()["id"]
    yonetici.post("/api/curriculum", json={
        "section_id": s, "subject_id": d, "teacher_id": o, "weekly_hours": 4,
    })
    pid = yonetici.post("/api/timetables", json={"name": "Kod"}).json()["id"]
    uret_ve_bekle(yonetici, pid)

    govde = yonetici.get(
        f"/api/timetables/{pid}/export/html?bakis=sube&duzen=carsaf"
    ).text
    assert ">MAT<" in govde and ">AY<" in govde


def test_carsaf_excel_tek_sayfa(yonetici: TestClient):
    import io
    from openpyxl import load_workbook

    _tanimlar(yonetici)
    pid = yonetici.post("/api/timetables", json={"name": "Çarşaf"}).json()["id"]
    uret_ve_bekle(yonetici, pid)

    r = yonetici.get(f"/api/timetables/{pid}/export/xlsx?bakis=sube&duzen=carsaf")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Çarşaf"]
    ws = wb["Çarşaf"]
    assert ws["A1"].value == "Şube"
    assert {ws.cell(row=r_, column=1).value for r_ in (3, 4)} == {"5-A", "5-B"}


def test_ayri_duzen_hala_sube_basina_sayfa_verir(yonetici: TestClient):
    import io
    from openpyxl import load_workbook

    _tanimlar(yonetici)
    pid = yonetici.post("/api/timetables", json={"name": "Ayrı"}).json()["id"]
    uret_ve_bekle(yonetici, pid)

    wb = load_workbook(io.BytesIO(
        yonetici.get(f"/api/timetables/{pid}/export/xlsx?bakis=sube").content
    ))
    assert sorted(wb.sheetnames) == ["5-A", "5-B"]


def test_gecersiz_duzen_reddedilir(yonetici: TestClient):
    pid = yonetici.post("/api/timetables", json={"name": "X"}).json()["id"]
    assert yonetici.get(
        f"/api/timetables/{pid}/export/html?duzen=yok"
    ).status_code == 422


def test_ogretmen_rengi_kaydedilir_ve_aktarilir(yonetici: TestClient):
    o = yonetici.post("/api/teachers", json={
        "full_name": "Renkli Öğretmen", "color": "#ef4444",
    }).json()
    assert o["color"] == "#ef4444"

    # Renk verilmezse varsayılan gelir.
    varsayilan = yonetici.post("/api/teachers", json={"full_name": "Sade"}).json()
    assert varsayilan["color"] == "#94a3b8"

    # Geçersiz renk reddedilir.
    assert yonetici.post("/api/teachers", json={
        "full_name": "Hatalı", "color": "kırmızı",
    }).status_code == 422


def test_renk_gecmis_donemden_aktarilirken_korunur(yonetici: TestClient):
    eski = yonetici.get("/api/terms").json()[0]["id"]
    kaynak = yonetici.post("/api/teachers", json={
        "full_name": "Renkli Öğretmen", "color": "#8b5cf6",
    }).json()["id"]

    yonetici.post("/api/terms", json={"name": "Yeni Dönem"})
    yonetici.post("/api/teachers/import", json={"term_id": eski, "ids": [kaynak]})
    assert yonetici.get("/api/teachers").json()[0]["color"] == "#8b5cf6"


def test_model_listesi_anahtar_yoksa_reddedilir(yonetici: TestClient):
    r = yonetici.post("/api/ai/models", json={})
    assert r.status_code == 400
    assert "API anahtarınızı girin" in r.text


def test_model_listesi_saglayiciyi_sorar(yonetici: TestClient, monkeypatch):
    """Kaydedilmemiş anahatla da sorgulanabilir; liste sıralı ve tekilleştirilmiş döner."""
    from app.ai import client as ai

    cagrilar = {}

    class SahteModel:
        def __init__(self, id): self.id = id

    class SahteListe:
        data = [SahteModel("gpt-4o-mini"), SahteModel("gpt-4o"), SahteModel("gpt-4o")]

    class SahteIstemci:
        def __init__(self, api_key, base_url):
            cagrilar["api_key"] = api_key
            cagrilar["base_url"] = base_url
            self.models = self

        def list(self):
            return SahteListe()

    monkeypatch.setattr(ai, "_istemci", lambda k, u: SahteIstemci(k, u))

    r = yonetici.post("/api/ai/models", json={
        "base_url": "https://ornek.local/v1", "api_key": "sk-deneme",
    })
    assert r.status_code == 200
    assert r.json()["models"] == ["gpt-4o", "gpt-4o-mini"]
    assert r.json()["source"] == "https://ornek.local/v1"
    assert cagrilar == {"api_key": "sk-deneme", "base_url": "https://ornek.local/v1"}


def test_model_listesi_kayitli_anahtari_kullanir(yonetici: TestClient, monkeypatch):
    from app.ai import client as ai

    gorulen = {}

    class Sahte:
        def __init__(self, api_key, base_url):
            gorulen["api_key"] = api_key
            self.models = self

        def list(self):
            class L: data = [type("M", (), {"id": "yerel-model"})()]
            return L()

    monkeypatch.setattr(ai, "_istemci", lambda k, u: Sahte(k, u))
    yonetici.put("/api/ai/settings", json={
        "enabled": True, "base_url": "http://localhost:11434/v1",
        "model": "llama3.1", "api_key": "gizli-anahtar",
    })

    r = yonetici.post("/api/ai/models", json={})
    assert r.status_code == 200
    assert r.json()["models"] == ["yerel-model"]
    assert gorulen["api_key"] == "gizli-anahtar"      # şifreli kayıttan çözüldü


def test_model_listesi_saglayici_hatasini_iletir(yonetici: TestClient, monkeypatch):
    from app.ai import client as ai

    def patla(k, u):
        raise RuntimeError("401 Unauthorized")

    monkeypatch.setattr(ai, "_istemci", patla)
    r = yonetici.post("/api/ai/models", json={"api_key": "sk-hatali"})
    assert r.status_code == 502
    assert "Adresi ve anahtarı kontrol edin" in r.text
    assert "401 Unauthorized" in r.text


def test_bos_model_listesi_hata_sayilir(yonetici: TestClient, monkeypatch):
    from app.ai import client as ai

    class Bos:
        def __init__(self, *a): self.models = self
        def list(self):
            class L: data = []
            return L()

    monkeypatch.setattr(ai, "_istemci", lambda k, u: Bos())
    r = yonetici.post("/api/ai/models", json={"api_key": "sk-x"})
    assert r.status_code == 502
    assert "boş bir model listesi" in r.text


def test_saglik_uclari(istemci: TestClient):
    assert istemci.get("/api/health").json() == {"status": "ok"}
    assert istemci.get("/api/health/db").json() == {"status": "ok"}


def test_veritabani_erisilemezse_503(istemci: TestClient, monkeypatch):
    """Veritabanı yoksa servis 'sağlıklı' demek yerine açıkça hata vermeli."""
    from app import main

    class Kopuk:
        def connect(self):
            raise OSError("Can't connect to MySQL server (110)")

    monkeypatch.setattr("app.db.engine", Kopuk())
    r = istemci.get("/api/health/db")
    assert r.status_code == 503
    assert "Veritabanına ulaşılamıyor" in r.text
    assert main is not None


def test_kapali_saatler_toplu_okunur(yonetici: TestClient):
    """Çarşaf görünümü kapalı saatleri tek istekte alır."""
    sube = yonetici.post("/api/sections", json={"name": "10-K"}).json()["id"]
    ogr = yonetici.post("/api/teachers", json={"full_name": "Kapalı Öğretmen"}).json()["id"]
    gunler = [g for g in yonetici.get("/api/timegrid").json() if g["is_active"]]
    sabah = [p["id"] for g in gunler for p in g["periods"] if p["index"] < 2]
    ilk = [p["id"] for g in gunler[:1] for p in g["periods"] if p["index"] == 0]

    yonetici.put(f"/api/sections/{sube}/availability", json={
        "cells": [{"period_id": pid, "state": "uygun_degil"} for pid in sabah]
    })
    yonetici.put(f"/api/teachers/{ogr}/availability", json={
        "cells": [{"period_id": pid, "state": "uygun_degil"} for pid in ilk]
    })

    kapali = yonetici.get("/api/availability/closed").json()
    assert sorted(kapali["sections"][str(sube)]) == sorted(sabah)
    assert sorted(kapali["teachers"][str(ogr)]) == sorted(ilk)


def test_kapali_saatler_baska_kuruma_sizmaz(yonetici: TestClient, istemci: TestClient):
    """Başka kurumun kapalı saatleri listeye karışmamalı."""
    sube = yonetici.post("/api/sections", json={"name": "11-K"}).json()["id"]
    gunler = [g for g in yonetici.get("/api/timegrid").json() if g["is_active"]]
    saat = [p["id"] for g in gunler[:1] for p in g["periods"] if p["index"] == 0]
    yonetici.put(f"/api/sections/{sube}/availability", json={
        "cells": [{"period_id": pid, "state": "uygun_degil"} for pid in saat]
    })

    istemci.post("/api/auth/register", json={
        "institution_name": "Öteki Okul", "term_name": "2026",
        "full_name": "Öteki Yönetici", "email": "oteki@ornek.com", "password": "sifre1234",
    })
    jeton = istemci.post("/api/auth/login", json={
        "email": "oteki@ornek.com", "password": "sifre1234",
    }).json()["access_token"]
    istemci.headers["Authorization"] = f"Bearer {jeton}"

    assert istemci.get("/api/availability/closed").json() == {
        "teachers": {}, "sections": {},
    }


def test_silinen_sube_kapali_saatleri_listelenmez(yonetici: TestClient):
    sube = yonetici.post("/api/sections", json={"name": "12-K"}).json()["id"]
    gunler = [g for g in yonetici.get("/api/timegrid").json() if g["is_active"]]
    saat = [p["id"] for g in gunler[:1] for p in g["periods"] if p["index"] == 0]
    yonetici.put(f"/api/sections/{sube}/availability", json={
        "cells": [{"period_id": pid, "state": "uygun_degil"} for pid in saat]
    })
    assert str(sube) in yonetici.get("/api/availability/closed").json()["sections"]

    yonetici.delete(f"/api/sections/{sube}")
    assert str(sube) not in yonetici.get("/api/availability/closed").json()["sections"]


def _carsaf_okul(c: TestClient) -> tuple[int, int]:
    """2 saatlik blok dersi ve sabahları kapalı bir şubesi olan küçük okul."""
    d = c.post("/api/subjects", json={
        "name": "Matematik", "short_code": "MAT",
    }).json()["id"]
    o = c.post("/api/teachers", json={
        "full_name": "Ayşe Yılmaz", "short_code": "AY",
    }).json()["id"]
    s = c.post("/api/sections", json={"name": "9-A"}).json()["id"]

    gunler = [g for g in c.get("/api/timegrid").json() if g["is_active"]]
    sabah = [p["id"] for p in gunler[0]["periods"] if p["index"] < 2]
    c.put(f"/api/sections/{s}/availability", json={
        "cells": [{"period_id": pid, "state": "uygun_degil"} for pid in sabah]
    })
    c.post("/api/curriculum", json={
        "section_id": s, "subject_id": d, "teacher_id": o,
        "weekly_hours": 2, "block_pattern": "2", "max_per_day": 2,
    })
    pid = c.post("/api/timetables", json={"name": "Çarşaf"}).json()["id"]
    uret_ve_bekle(c, pid)
    return pid, len(sabah)


def test_carsaf_ciktisi_ardisik_saatleri_birlestirir(yonetici: TestClient):
    """İki saatlik blok tek hücrede çıkmalı — ekrandaki çarşafla aynı düzen."""
    pid, _ = _carsaf_okul(yonetici)
    html = yonetici.get(
        f"/api/timetables/{pid}/export/html?bakis=sube&duzen=carsaf"
    ).text
    assert 'colspan="2"' in html
    # Blok tek hücrede olduğu için kısaltma bir kez yazılır.
    assert html.count('<span class="ders">MAT</span>') == 1


def test_carsaf_ciktisi_kapali_saatleri_isaretler(yonetici: TestClient):
    pid, kapali_sayisi = _carsaf_okul(yonetici)
    html = yonetici.get(
        f"/api/timetables/{pid}/export/html?bakis=sube&duzen=carsaf"
    ).text
    assert html.count('class="kpl') == kapali_sayisi


def test_ayri_sayfa_ciktisi_birlestirmez(yonetici: TestClient):
    """Ayrı sayfa düzeni saat saat okunur; orada birleştirme yapılmaz."""
    pid, _ = _carsaf_okul(yonetici)
    html = yonetici.get(
        f"/api/timetables/{pid}/export/html?bakis=sube&duzen=ayri"
    ).text
    assert "colspan" not in html


# --- Öğle arası ve öğretmenin gün sınırı ---

def _ogle_arasi_ekle(c: TestClient, sira: int = 4) -> list[dict]:
    """Izgaradaki bir saati öğle arasına çevirir; sonrası öğleden sonra olur."""
    gunler = c.get("/api/timegrid").json()
    for g in gunler:
        for p in g["periods"]:
            p["is_lunch"] = p["index"] == sira
    r = c.put("/api/timegrid", json=gunler)
    assert r.status_code == 200, r.text
    return r.json()


def test_ogle_arasi_kaydedilir_ve_teneffus_olur(yonetici: TestClient):
    gunler = _ogle_arasi_ekle(yonetici)
    ogle = [p for g in gunler for p in g["periods"] if p["is_lunch"]]
    assert len(ogle) == len([g for g in gunler if g["periods"]])
    # Öğle arasına ders konmaz: işaretlenince teneffüs de olur.
    assert all(p["is_break"] for p in ogle)


def test_gunde_iki_ogle_arasi_reddedilir(yonetici: TestClient):
    gunler = yonetici.get("/api/timegrid").json()
    for p in gunler[0]["periods"][:2]:
        p["is_lunch"] = True
    r = yonetici.put("/api/timegrid", json=gunler)
    assert r.status_code == 422
    assert "öğle arası" in r.text


def test_gun_siniri_yarim_kabul_eder(yonetici: TestClient):
    r = yonetici.post("/api/teachers", json={
        "full_name": "Yarım Günlü", "max_days": 4.5,
    })
    assert r.status_code == 201, r.text
    assert r.json()["max_days"] == 4.5
    assert yonetici.get("/api/teachers").json()[0]["max_days"] == 4.5


def test_gun_siniri_ceyrek_kabul_etmez(yonetici: TestClient):
    r = yonetici.post("/api/teachers", json={
        "full_name": "Çeyrek Günlü", "max_days": 4.25,
    })
    assert r.status_code == 422
    assert "yarım" in r.text


def test_gun_siniri_bos_birakilabilir(yonetici: TestClient):
    r = yonetici.post("/api/teachers", json={"full_name": "Sınırsız"})
    assert r.status_code == 201
    assert r.json()["max_days"] is None


def test_gun_siniri_programa_uygulanir(yonetici: TestClient):
    """4 gün anlaşması olan öğretmen 5 güne yayılmaz."""
    _ogle_arasi_ekle(yonetici)
    d = yonetici.post("/api/subjects", json={"name": "Matematik"}).json()["id"]
    o = yonetici.post("/api/teachers", json={
        "full_name": "Dört Günlü", "max_days": 4,
    }).json()["id"]
    s = yonetici.post("/api/sections", json={"name": "9-A"}).json()["id"]
    yonetici.post("/api/curriculum", json={
        "section_id": s, "subject_id": d, "teacher_id": o,
        "weekly_hours": 8, "max_per_day": 2,
    })

    pid = yonetici.post("/api/timetables", json={"name": "Gün sınırı"}).json()["id"]
    deneme = uret_ve_bekle(yonetici, pid)
    assert deneme["status"] == "basarili", deneme["report"]

    hucreler = yonetici.get(f"/api/timetables/{pid}/grid").json()["cells"]
    assert len({h["day_index"] for h in hucreler}) <= 4


def test_gun_siniri_asilirsa_uyari_cikar(yonetici: TestClient):
    """Sınıra sığmayan yük: program üretilir, aşım uyarı olarak listelenir."""
    d = yonetici.post("/api/subjects", json={"name": "Yoğun"}).json()["id"]
    o = yonetici.post("/api/teachers", json={
        "full_name": "İki Günlü", "max_days": 2,
    }).json()["id"]
    s = yonetici.post("/api/sections", json={"name": "9-Y"}).json()["id"]
    # 2 günde en çok 16 saat var; 24 saat isteniyor. Desen günü dolduran
    # bloklar verir: desensiz 24 saat tek saatlere açılır ve kural 7 (aynı
    # dersin saatleri bitişik olamaz) yükü hiçbir sınırla yerleştiremez.
    yonetici.post("/api/curriculum", json={
        "section_id": s, "subject_id": d, "teacher_id": o,
        "weekly_hours": 24, "block_pattern": "8+8+8", "max_per_day": 8,
    })

    pid = yonetici.post("/api/timetables", json={"name": "Aşım"}).json()["id"]
    uret_ve_bekle(yonetici, pid, saniye=120.0)

    uyarilar = yonetici.get(f"/api/timetables/{pid}/warnings").json()
    gun = [u for u in uyarilar if u["tur"] == "gun_siniri"]
    assert gun, uyarilar
    assert gun[0]["ogretmen"] == "İki Günlü"
    assert gun[0]["sinir"] == 4          # yarım gün biriminde: 2 gün
    assert gun[0]["konan"] > 4


def test_gun_siniri_uyarisi_gizlenebilir(yonetici: TestClient):
    d = yonetici.post("/api/subjects", json={"name": "Yoğun2"}).json()["id"]
    o = yonetici.post("/api/teachers", json={
        "full_name": "Tek Günlü", "max_days": 1,
    }).json()["id"]
    s = yonetici.post("/api/sections", json={"name": "9-Z"}).json()["id"]
    yonetici.post("/api/curriculum", json={
        "section_id": s, "subject_id": d, "teacher_id": o,
        "weekly_hours": 16, "block_pattern": "8+8", "max_per_day": 8,
    })
    pid = yonetici.post("/api/timetables", json={"name": "Gizle"}).json()["id"]
    uret_ve_bekle(yonetici, pid, saniye=120.0)

    uyari = next(u for u in yonetici.get(f"/api/timetables/{pid}/warnings").json()
                 if u["tur"] == "gun_siniri")
    sonra = yonetici.post(f"/api/timetables/{pid}/warnings/ignore",
                          json={"key": uyari["key"]}).json()
    assert next(u for u in sonra if u["key"] == uyari["key"])["ignored"] is True


def test_gun_siniri_baska_doneme_aktarilir(yonetici: TestClient):
    yonetici.post("/api/teachers", json={"full_name": "Taşınan", "max_days": 3.5})
    ilk = yonetici.get("/api/terms").json()[0]["id"]
    yeni = yonetici.post("/api/terms", json={"name": "Sonraki"}).json()["id"]
    yonetici.post(f"/api/terms/{yeni}/activate")

    aday = yonetici.get(f"/api/teachers/import/{ilk}").json()
    hedef = next(t for t in aday if t["full_name"] == "Taşınan")
    yonetici.post("/api/teachers/import", json={"term_id": ilk, "ids": [hedef["id"]]})

    tasinan = next(t for t in yonetici.get("/api/teachers").json()
                   if t["full_name"] == "Taşınan")
    assert tasinan["max_days"] == 3.5


# --- Ders saatlerini yeniden sıralama ---

def _gun_ve_saatler(c: TestClient) -> tuple[dict, list[dict]]:
    gun = c.get("/api/timegrid").json()[0]
    return gun, sorted(gun["periods"], key=lambda p: p["index"])


def test_saatler_yeniden_siralanabilir(yonetici: TestClient):
    """Son satır başa çekilince sıra değişir, kimlikler korunur."""
    gunler = yonetici.get("/api/timegrid").json()
    saatler = sorted(gunler[0]["periods"], key=lambda p: p["index"])
    tasinan = saatler[-1]

    yeni_sira = [tasinan] + saatler[:-1]
    gunler[0]["periods"] = [
        {**p, "index": i} for i, p in enumerate(yeni_sira)
    ]
    r = yonetici.put("/api/timegrid", json=gunler)
    assert r.status_code == 200, r.text

    _, sonra = _gun_ve_saatler(yonetici)
    assert sonra[0]["id"] == tasinan["id"]
    assert [p["id"] for p in sonra] == [p["id"] for p in yeni_sira]


def test_siralama_musaitligi_tasinan_satirla_gotur(yonetici: TestClient):
    """Araya teneffüs eklenince işaret, kaydığı satırın peşinden gitmeli.

    Sıraya göre eşleştirilseydi "3. ders"in kapalı işareti yerinde kalır,
    araya giren teneffüsün üstüne yapışırdı.
    """
    ogretmen = yonetici.post("/api/teachers", json={"full_name": "Sıra Testi"}).json()
    gunler = yonetici.get("/api/timegrid").json()
    saatler = sorted(gunler[0]["periods"], key=lambda p: p["index"])
    ucuncu = saatler[2]
    yonetici.put(f"/api/teachers/{ogretmen['id']}/availability", json={
        "cells": [{"period_id": ucuncu["id"], "state": "uygun_degil"}]
    })

    # 1. dersten sonra bir teneffüs; sonraki satırların hepsi bir kayıyor.
    yeni = {"id": None, "index": 0, "name": "Teneffüs", "start_time": None,
            "end_time": None, "is_break": True, "is_lunch": False}
    yeni_sira = saatler[:1] + [yeni] + saatler[1:]
    gunler[0]["periods"] = [{**p, "index": i} for i, p in enumerate(yeni_sira)]
    assert yonetici.put("/api/timegrid", json=gunler).status_code == 200

    _, sonra = _gun_ve_saatler(yonetici)
    assert sonra[1]["is_break"] is True          # araya teneffüs girdi
    assert sonra[3]["id"] == ucuncu["id"]        # "3. ders" bir aşağı kaydı

    kapali = yonetici.get(f"/api/teachers/{ogretmen['id']}/availability").json()
    assert [h["period_id"] for h in kapali] == [ucuncu["id"]]


def test_ortadaki_saat_silinebilir(yonetici: TestClient):
    """Gönderilmeyen satır, müsaitlik işaretleriyle birlikte silinir."""
    ogretmen = yonetici.post("/api/teachers", json={"full_name": "Silme Testi"}).json()
    gunler = yonetici.get("/api/timegrid").json()
    saatler = sorted(gunler[0]["periods"], key=lambda p: p["index"])
    silinecek = saatler[1]
    yonetici.put(f"/api/teachers/{ogretmen['id']}/availability", json={
        "cells": [{"period_id": silinecek["id"], "state": "uygun_degil"}]
    })

    kalan = [p for p in saatler if p["id"] != silinecek["id"]]
    gunler[0]["periods"] = [{**p, "index": i} for i, p in enumerate(kalan)]
    assert yonetici.put("/api/timegrid", json=gunler).status_code == 200

    _, sonra = _gun_ve_saatler(yonetici)
    assert silinecek["id"] not in [p["id"] for p in sonra]
    assert yonetici.get(f"/api/teachers/{ogretmen['id']}/availability").json() == []


def test_baska_gunun_saati_gonderilemez(yonetici: TestClient):
    gunler = yonetici.get("/api/timegrid").json()
    yabanci = sorted(gunler[1]["periods"], key=lambda p: p["index"])[0]
    gunler[0]["periods"] = [
        {**p, "id": yabanci["id"] if i == 0 else p["id"]}
        for i, p in enumerate(sorted(gunler[0]["periods"], key=lambda p: p["index"]))
    ]
    r = yonetici.put("/api/timegrid", json=gunler)
    assert r.status_code == 422
    assert "ait olmayan" in r.text


def test_ayni_sira_iki_kez_gonderilemez(yonetici: TestClient):
    gunler = yonetici.get("/api/timegrid").json()
    saatler = sorted(gunler[0]["periods"], key=lambda p: p["index"])
    gunler[0]["periods"] = [{**saatler[0], "index": 0}, {**saatler[1], "index": 0}]
    r = yonetici.put("/api/timegrid", json=gunler)
    assert r.status_code == 422
    assert "sıra numarası" in r.text


def test_kimliksiz_gonderim_eskisi_gibi_calisir(yonetici: TestClient):
    """Kimlik göndermeyen istemci sıraya göre eşleştirilmeye devam eder."""
    gunler = yonetici.get("/api/timegrid").json()
    saatler = sorted(gunler[0]["periods"], key=lambda p: p["index"])
    kimlikler = [p["id"] for p in saatler]
    gunler[0]["periods"] = [
        {"index": p["index"], "name": f"Ders {p['index'] + 1}", "start_time": None,
         "end_time": None, "is_break": False, "is_lunch": False}
        for p in saatler
    ]
    assert yonetici.put("/api/timegrid", json=gunler).status_code == 200

    _, sonra = _gun_ve_saatler(yonetici)
    assert [p["id"] for p in sonra] == kimlikler      # kayıtlar korundu
    assert sonra[0]["name"] == "Ders 1"
